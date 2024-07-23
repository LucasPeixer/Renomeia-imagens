import openpyxl
import os

# Caminho para a sua planilha Excel
caminho_planilha = 'C:\\Users\\LUCAS PEIXER\\Downloads\\blumenau\\form.xlsx'

# Caminho para a pasta com as imagens
caminho_imagens = 'C:\\Users\\LUCAS PEIXER\\Downloads\\blumenau\\form_1'

# Carrega a planilha Excel
workbook = openpyxl.load_workbook(caminho_planilha)
sheet = workbook.active

# Itera sobre as linhas da tabela (excluindo a primeira linha, que provavelmente é o cabeçalho)
for linha in range(2, sheet.max_row + 1):
    # Obtém o ID da linha atual
    id_linha = sheet.cell(row=linha, column=15).value  # Pegando a info da coluna ID

    # Obtém o nome da imagem da coluna "imagens"
    nomes_imagens = sheet.cell(row=linha, column=93).value  # Pegando a info da coluna imagens
    nomes_imagens = nomes_imagens.split(',')  # Separa os nomes por vírgula

    # Itera sobre os nomes das imagens
    for i, nome_imagem in enumerate(nomes_imagens):
        nome_imagem = nome_imagem.strip()  # Remove espaços em branco antes e depois do nome
        # Procura o arquivo de imagem com o nome correspondente
        for arquivo in os.listdir(caminho_imagens):
            if nome_imagem in arquivo and arquivo.endswith('.jpg'):
                # Renomeia a imagem com o ID da linha
                novo_nome = f'{id_linha}_{i+1}.jpg'  # Adiciona um número sequencial para cada imagem
                os.rename(os.path.join(caminho_imagens, arquivo), os.path.join(caminho_imagens, novo_nome))
                print('Feito imagem {nome_imagem}')
                break  


